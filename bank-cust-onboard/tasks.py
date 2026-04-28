from pathlib import Path
import re

from openpyxl import load_workbook
from robocorp.tasks import get_output_dir, task, workitems


@task
def producer():
    """Split Excel rows into multiple output Work Items for the next step."""
    print("PRODUCER STARTED")
    output = get_output_dir() or Path("output")
    filename = "orders.xlsx"

    for item in workitems.inputs:
        path = item.get_file(filename, output / filename)

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            print("Rows found:", 0)
            continue

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        header_index = {name: idx for idx, name in enumerate(headers)}
        print("Rows found:", len(rows) - 1)

        for row in rows[1:]:
            payload = {
                "Name": row[header_index["Name"]] if "Name" in header_index and header_index["Name"] < len(row) else None,
                "Zip": row[header_index["Zip"]] if "Zip" in header_index and header_index["Zip"] < len(row) else None,
                "Product": row[header_index["Item"]] if "Item" in header_index and header_index["Item"] < len(row) else None,
            }
            workitems.outputs.create(payload)


@task
def consumer():
    print("CONSUMER STARTED")
    zip_code_re = r"^\d{5}(-\d{4})?$"

    for item in workitems.inputs:
        try:
            name = item.payload["Name"]
            zipcode = item.payload["Zip"]
            product = item.payload["Product"]

            print(f"Processing order: {name}, {zipcode}, {product}")

            if not re.match(zip_code_re, str(zipcode)):
                raise AssertionError("Invalid ZIP code")

            item.done()

        except AssertionError as err:
            item.fail("BUSINESS", code="INVALID_ORDER", message=str(err))

        except KeyError as err:
            item.fail("APPLICATION", code="MISSING_FIELD", message=str(err))
